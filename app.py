from flask import Flask, render_template, request, redirect, url_for, flash, session
from datetime import datetime, date, timedelta
import os
import secrets
from flask_wtf.csrf import CSRFProtect, generate_csrf
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

from models import db, Medicine, Customer, Sale
from sqlalchemy import func
from io import BytesIO
from flask import send_file

# Configuration
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def create_app():
    from models import Admin
    from werkzeug.security import generate_password_hash, check_password_hash
    from functools import wraps

    app = Flask(__name__)
    # Use PostgreSQL on Railway, SQLite locally
    database_url = os.environ.get('DATABASE_URL')
    if database_url:
        # Fix PostgreSQL URL scheme for SQLAlchemy
        if database_url.startswith('postgres://'):
            database_url = database_url.replace('postgres://', 'postgresql://', 1)
        app.config['SQLALCHEMY_DATABASE_URI'] = database_url
    else:
        # Local development with SQLite
        DB_PATH = os.path.join(BASE_DIR, 'pharmacy.db')
        app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_PATH}'
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.secret_key = os.environ.get('SECRET_KEY', 'dev-key-for-demo')

    # Initialize database object from models.py
    db.init_app(app)

    # Initialize CSRF protection
    csrf = CSRFProtect(app)

    # Make generate_csrf available in all templates
    app.jinja_env.globals['generate_csrf'] = generate_csrf

    # Initialize rate limiter
    limiter = Limiter(
        key_func=get_remote_address,
        app=app,
        default_limits=["200 per day", "50 per hour"]
    )

    with app.app_context():
        try:
            # Create tables if they don't exist
            db.create_all()
        except Exception as e:
            print(f"Warning: Could not create database tables: {e}")
            # Continue anyway - tables might already exist

    # Add context processor to inject current date and time in 12-hour format
    @app.context_processor
    def inject_now():
        now = datetime.now()
        hour = now.strftime('%I')
        if hour.startswith('0'):
            hour = hour[1:]
        current_time = f"{hour}{now.strftime(':%M:%S %p')}"
        current_date = now.strftime('%B %d, %Y')
        return dict(current_time=current_time, current_date=current_date)

    # Add context processor for CSRF token
    @app.context_processor
    def inject_csrf():
        return dict(generate_csrf=generate_csrf)

    # Admin decorator
    def admin_required(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not session.get('is_admin'):
                flash('Admin login required.', 'warning')
                return redirect(url_for('admin_login'))
            return f(*args, **kwargs)
        return decorated_function

    # Admin login route
    @app.route('/admin/login', methods=['GET', 'POST'])
    @limiter.limit("5 per hour", methods=["POST"])
    def admin_login():
        error = None
        if request.method == 'POST':
            email = request.form.get('email', '').strip().lower()
            password = request.form.get('password', '')
            admin = Admin.query.filter_by(email=email).first()
            if admin and admin.check_password(password):
                session['is_admin'] = True
                session['admin_email'] = email
                flash('Logged in as admin.', 'success')
                return redirect(url_for('medicines'))
            else:
                error = 'Invalid email or password.'
        if Admin.query.first() is None:
            return render_template('admin_login.html', error=error, show_register=True)
        return render_template('admin_login.html', error=error, show_register=False)

    # Admin logout route
    @app.route('/admin/logout')
    def admin_logout():
        session.pop('is_admin', None)
        session.pop('admin_email', None)
        flash('Logged out.', 'info')
        return redirect(url_for('index'))

    # Admin registration route
    @app.route('/admin/register', methods=['GET', 'POST'])
    @limiter.limit("3 per hour", methods=["POST"])
    def admin_register():
        if Admin.query.first():
            flash('Admin already registered. Only one admin allowed for MVP.', 'info')
            return redirect(url_for('admin_login'))
        error = None
        if request.method == 'POST':
            email = request.form.get('email', '').strip().lower()
            phone = request.form.get('phone', '').strip()
            password = request.form.get('password', '')
            confirm = request.form.get('confirm', '')
            # Basic validation
            if not email or not phone or not password or not confirm:
                error = 'All fields are required.'
            elif password != confirm:
                error = 'Passwords do not match.'
            elif len(password) < 8 or not any(c.isalpha() for c in password) or not any(not c.isalnum() for c in password):
                error = 'Password must be at least 8 characters, include a letter and a special character.'
            elif Admin.query.filter_by(email=email).first():
                error = 'Email already registered.'
            else:
                admin = Admin(email=email, phone=phone)
                admin.set_password(password)
                db.session.add(admin)
                db.session.commit()
                flash('Admin registered. Please log in.', 'success')
                return redirect(url_for('admin_login'))
        return render_template('admin_register.html', error=error)

    # Admin password reset routes
    @app.route('/admin/forgot-password', methods=['GET', 'POST'])
    def forgot_password():
        error = None
        success = None
        if request.method == 'POST':
            email_or_phone = request.form.get('email_or_phone', '').strip().lower()
            # Search by email or phone
            admin = Admin.query.filter(
                (Admin.email == email_or_phone) | (Admin.phone == email_or_phone)
            ).first()
            if admin:
                # Generate a reset token and store it in session (demo only)
                reset_token = secrets.token_urlsafe(32)
                session['reset_token'] = reset_token
                session['reset_email'] = admin.email
                success = f'Reset link sent! Redirecting...'
                flash(success, 'success')
                return redirect(url_for('reset_password', token=reset_token))
            else:
                error = 'No admin found with that email or phone number.'
        return render_template('forgot_password.html', error=error)

    @app.route('/admin/reset-password/<token>', methods=['GET', 'POST'])
    @limiter.limit("5 per hour", methods=["POST"])
    def reset_password(token):
        # Verify token is in session
        if session.get('reset_token') != token:
            flash('Invalid or expired reset link.', 'danger')
            return redirect(url_for('admin_login'))
        
        error = None
        if request.method == 'POST':
            new_password = request.form.get('password', '')
            confirm = request.form.get('confirm', '')
            
            # Validation
            if not new_password or not confirm:
                error = 'Both fields are required.'
            elif new_password != confirm:
                error = 'Passwords do not match.'
            elif len(new_password) < 8 or not any(c.isalpha() for c in new_password) or not any(not c.isalnum() for c in new_password):
                error = 'Password must be at least 8 characters, include a letter and a special character.'
            else:
                # Update password
                admin_email = session.get('reset_email')
                admin = Admin.query.filter_by(email=admin_email).first()
                if admin:
                    admin.set_password(new_password)
                    db.session.commit()
                    # Clear reset session
                    session.pop('reset_token', None)
                    session.pop('reset_email', None)
                    flash('Password reset successfully! Please log in.', 'success')
                    return redirect(url_for('admin_login'))
                else:
                    error = 'Admin not found.'
        
        return render_template('reset_password.html', token=token, error=error)

    # Basic role selection landing page
    @app.route('/')
    def index():
        return render_template('index.html')

    # ---------- DASHBOARD ROUTE ----------
    @app.route('/dashboard')
    def dashboard():
        """Dashboard with charts and analytics"""
        try:
            # Sales trend data (last 7 days)
            today = date.today()
            seven_days_ago = today - timedelta(days=7)
            sales_by_date = db.session.query(
                func.date(Sale.timestamp).label('date'),
                func.sum(Sale.total_price).label('total')
            ).filter(Sale.timestamp >= seven_days_ago).group_by(func.date(Sale.timestamp)).all()
            
            sales_dates = [str(s[0]) for s in sales_by_date]
            sales_amounts = [float(s[1]) if s[1] else 0 for s in sales_by_date]
            
            # Stock levels
            medicines = Medicine.query.all()
            stock_labels = [m.name for m in medicines[:10]]  # Top 10
            stock_quantities = [m.quantity for m in medicines[:10]]
            
            # Expiry alerts
            today = date.today()
            thirty_days = today + timedelta(days=30)
            expiring_soon = Medicine.query.filter(
                Medicine.expiry_date.isnot(None),
                Medicine.expiry_date <= thirty_days,
                Medicine.expiry_date > today
            ).all()
            
            expired = Medicine.query.filter(
                Medicine.expiry_date.isnot(None),
                Medicine.expiry_date <= today
            ).all()
            
            # Statistics
            total_medicines = Medicine.query.count()
            total_stock = db.session.query(func.sum(Medicine.quantity)).scalar() or 0
            total_sales = db.session.query(func.sum(Sale.total_price)).scalar() or 0
            
            return render_template('dashboard.html',
                                 sales_dates=sales_dates,
                                 sales_amounts=sales_amounts,
                                 stock_labels=stock_labels,
                                 stock_quantities=stock_quantities,
                                 expiring_soon=expiring_soon,
                                 expired=expired,
                                 total_medicines=total_medicines,
                                 total_stock=total_stock,
                                 total_sales=total_sales)
        except Exception as e:
            flash(f'Dashboard error: {str(e)}', 'danger')
            return render_template('dashboard.html',
                                 sales_dates=[],
                                 sales_amounts=[],
                                 stock_labels=[],
                                 stock_quantities=[],
                                 expiring_soon=[],
                                 expired=[],
                                 total_medicines=0,
                                 total_stock=0,
                                 total_sales=0)

    # ---------- Medicine Inventory Routes ----------
    @app.route('/medicines')
    def medicines():
        medicines = Medicine.query.order_by(Medicine.name).all()
        low_stock_threshold = 5
        return render_template('medicines.html', medicines=medicines, low_stock_threshold=low_stock_threshold, today=date.today())

    @app.route('/medicines/add', methods=['GET', 'POST'])
    @admin_required
    def add_medicine():
        if request.method == 'POST':
            try:
                name = request.form['name']
                brand = request.form.get('brand')
                price = float(request.form['price'])
                quantity = int(request.form['quantity'])
                expiry = request.form.get('expiry_date') or None
                expiry_date = datetime.strptime(expiry, '%Y-%m-%d').date() if expiry else None
                category = request.form.get('category')
                description = request.form.get('description') or None

                med = Medicine(name=name, brand=brand, price=price, quantity=quantity, expiry_date=expiry_date, category=category, description=description)
                db.session.add(med)
                db.session.commit()
                flash('Medicine added successfully.', 'success')
                return redirect(url_for('medicines'))
            except ValueError as e:
                flash(f'Invalid input: Please check your entries. {str(e)}', 'danger')
                return render_template('add_medicine.html')
            except Exception as e:
                db.session.rollback()
                flash(f'Error adding medicine: {str(e)}', 'danger')
                return render_template('add_medicine.html')

        return render_template('add_medicine.html')

    @app.route('/medicines/update/<int:med_id>', methods=['GET', 'POST'])
    @admin_required
    def update_medicine(med_id):
        med = Medicine.query.get_or_404(med_id)
        if request.method == 'POST':
            try:
                # Update all editable fields
                med.name = request.form.get('name') or med.name
                med.brand = request.form.get('brand') or med.brand
                med.category = request.form.get('category') or med.category
                med.price = float(request.form.get('price') or med.price)
                med.quantity = int(request.form.get('quantity') or med.quantity)
                expiry = request.form.get('expiry_date') or None
                med.expiry_date = datetime.strptime(expiry, '%Y-%m-%d').date() if expiry else None
                med.description = request.form.get('description') or med.description
                db.session.commit()
                flash('Medicine updated.', 'success')
                return redirect(url_for('medicines'))
            except ValueError as e:
                db.session.rollback()
                flash(f'Invalid input: Please check your entries. {str(e)}', 'danger')
                return render_template('update_medicine.html', med=med)
            except Exception as e:
                db.session.rollback()
                flash(f'Error updating medicine: {str(e)}', 'danger')
                return render_template('update_medicine.html', med=med)
        return render_template('update_medicine.html', med=med)

    @app.route('/medicines/delete/<int:med_id>', methods=['POST'])
    @admin_required
    def delete_medicine(med_id):
        med = Medicine.query.get_or_404(med_id)
        db.session.delete(med)
        db.session.commit()
        flash('Medicine deleted.', 'info')
        return redirect(url_for('medicines'))

    # ---------- Sales & Billing Routes ----------
    @app.route('/sales/new', methods=['GET', 'POST'])
    def new_sale():
        medicines = Medicine.query.filter(Medicine.quantity > 0).order_by(Medicine.name).all()
        customers = Customer.query.order_by(Customer.name).all()
        
        if request.method == 'POST':
            # Handle both form submissions and JSON API requests (from offline mode)
            is_json = request.is_json
            try:
                if is_json:
                    data = request.get_json()
                    med_id = int(data.get('medicine_id'))
                    qty = int(data.get('quantity'))
                    customer_id = data.get('customer_id')
                else:
                    med_id = int(request.form['medicine_id'])
                    qty = int(request.form['quantity'])
                    customer_id = request.form.get('customer_id') or None

                med = Medicine.query.get_or_404(med_id)
                if qty <= 0:
                    msg = 'Quantity must be positive.'
                    if is_json:
                        return jsonify({'error': msg}), 400
                    flash(msg, 'danger')
                    return redirect(url_for('new_sale'))
                if med.quantity < qty:
                    msg = 'Not enough stock for that medicine.'
                    if is_json:
                        return jsonify({'error': msg}), 400
                    flash(msg, 'danger')
                    return redirect(url_for('new_sale'))

                price_per_unit = med.price
                total_price = round(price_per_unit * qty, 2)

                # Reduce stock
                med.quantity -= qty

                sale = Sale(medicine=med, quantity=qty, price_per_unit=price_per_unit, total_price=total_price, customer_id=int(customer_id) if customer_id else None)
                db.session.add(sale)
                db.session.commit()
                
                if is_json:
                    return jsonify({'success': True, 'sale_id': sale.id, 'total': total_price}), 201
                
                flash('Sale recorded.', 'success')
                return redirect(url_for('receipt', sale_id=sale.id))
            except ValueError:
                msg = 'Invalid medicine or quantity.'
                if is_json:
                    return jsonify({'error': msg}), 400
                flash(msg, 'danger')
                return redirect(url_for('new_sale'))
            except Exception as e:
                db.session.rollback()
                msg = f'Error recording sale: {str(e)}'
                if is_json:
                    return jsonify({'error': msg}), 500
                flash(msg, 'danger')
                return redirect(url_for('new_sale'))

        return render_template('new_sale.html', medicines=medicines, customers=customers)

    @app.route('/sales/receipt/<int:sale_id>')
    def receipt(sale_id):
        sale = Sale.query.get_or_404(sale_id)
        return render_template('receipt.html', sale=sale)

    # ---------- Customers ----------
    @app.route('/customers', methods=['GET', 'POST'])
    def customers():
        if request.method == 'POST':
            try:
                name = request.form.get('name', '').strip()
                if not name:
                    flash('Customer name is required.', 'danger')
                    return redirect(url_for('customers'))
                phone = request.form.get('phone', '').strip()
                cust = Customer(name=name, phone=phone)
                db.session.add(cust)
                db.session.commit()
                flash('Customer added.', 'success')
                return redirect(url_for('customers'))
            except Exception as e:
                db.session.rollback()
                flash(f'Error adding customer: {str(e)}', 'danger')
                return redirect(url_for('customers'))
        customers = Customer.query.order_by(Customer.name).all()
        return render_template('customers.html', customers=customers)

    # ---------- Reports (MVP level) ----------
    @app.route('/reports')
    @admin_required
    def reports():
        """Main reports page with advanced analytics"""
        today = date.today()
        
        # ===== DAILY SALES =====
        start = datetime(today.year, today.month, today.day)
        end = datetime(today.year, today.month, today.day, 23, 59, 59)
        daily_sales = Sale.query.filter(Sale.timestamp >= start, Sale.timestamp <= end).order_by(Sale.timestamp.desc()).all()
        total_daily = sum(s.total_price for s in daily_sales)

        # ===== WEEKLY SALES =====
        seven_days_ago = today - timedelta(days=7)
        week_start = datetime(seven_days_ago.year, seven_days_ago.month, seven_days_ago.day)
        weekly_sales = db.session.query(
            func.date(Sale.timestamp).label('date'),
            func.sum(Sale.total_price).label('total'),
            func.count(Sale.id).label('count')
        ).filter(Sale.timestamp >= week_start).group_by(func.date(Sale.timestamp)).order_by('date').all()
        total_weekly = sum(w[1] for w in weekly_sales) if weekly_sales else 0

        # ===== MONTHLY SALES =====
        thirty_days_ago = today - timedelta(days=30)
        month_start = datetime(thirty_days_ago.year, thirty_days_ago.month, thirty_days_ago.day)
        monthly_sales = db.session.query(
            func.date(Sale.timestamp).label('date'),
            func.sum(Sale.total_price).label('total'),
            func.count(Sale.id).label('count')
        ).filter(Sale.timestamp >= month_start).group_by(func.date(Sale.timestamp)).order_by('date').all()
        total_monthly = sum(m[1] for m in monthly_sales) if monthly_sales else 0

        # ===== BEST-SELLING MEDICINES =====
        best_sellers = db.session.query(
            Medicine.id,
            Medicine.name,
            Medicine.price,
            func.sum(Sale.quantity).label('total_qty'),
            func.sum(Sale.total_price).label('total_revenue')
        ).join(Sale).group_by(Medicine.id, Medicine.name, Medicine.price).order_by(func.sum(Sale.quantity).desc()).limit(10).all()

        # ===== EXPIRED STOCK REPORT =====
        today = date.today()
        expired_items = Medicine.query.filter(
            Medicine.expiry_date.isnot(None),
            Medicine.expiry_date <= today
        ).order_by(Medicine.expiry_date.desc()).all()
        
        expiring_soon = Medicine.query.filter(
            Medicine.expiry_date.isnot(None),
            Medicine.expiry_date > today,
            Medicine.expiry_date <= today + timedelta(days=30)
        ).order_by(Medicine.expiry_date).all()

        # ===== PROFIT & LOSS =====
        all_sales = Sale.query.all()
        total_revenue = sum(s.total_price for s in all_sales)
        
        # Calculate total cost (price * quantity for all sales)
        total_cost = 0
        for sale in all_sales:
            # Assuming cost is calculated from medicine base price
            # For a real system, you'd store cost separately
            total_cost += (sale.medicine.price * sale.quantity) * 0.6  # Assuming 40% markup (60% cost)
        
        total_profit = total_revenue - total_cost
        profit_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0

        # Total sales summary (all time)
        total_all = sum(s.total_price for s in all_sales)

        # Stock report
        medicines = Medicine.query.order_by(Medicine.name).all()

        return render_template('reports.html',
                             daily_sales=daily_sales,
                             total_daily=total_daily,
                             total_weekly=total_weekly,
                             weekly_sales=weekly_sales,
                             total_monthly=total_monthly,
                             monthly_sales=monthly_sales,
                             total_all=total_all,
                             medicines=medicines,
                             best_sellers=best_sellers,
                             expired_items=expired_items,
                             expiring_soon=expiring_soon,
                             total_revenue=total_revenue,
                             total_cost=total_cost,
                             total_profit=total_profit,
                             profit_margin=profit_margin,
                             today=today)

    @app.route('/sales/search')
    @admin_required
    def search_sales():
        # Dedicated sales search endpoint, separate from reports
        q = request.args.get('q', '').strip()
        from_date = request.args.get('from_date')
        to_date = request.args.get('to_date')

        all_sales_q = Sale.query
        # Filter by date range if provided
        try:
            if from_date:
                fd = datetime.strptime(from_date, '%Y-%m-%d')
                all_sales_q = all_sales_q.filter(Sale.timestamp >= fd)
            if to_date:
                td = datetime.strptime(to_date, '%Y-%m-%d')
                td_end = datetime(td.year, td.month, td.day, 23, 59, 59)
                all_sales_q = all_sales_q.filter(Sale.timestamp <= td_end)
        except Exception:
            pass

        if q:
            if q.isdigit():
                all_sales_q = all_sales_q.filter(Sale.id == int(q))
            else:
                all_sales_q = all_sales_q.join(Sale.medicine).outerjoin(Sale.customer).filter(
                    (Medicine.name.ilike(f"%{q}%")) |
                    (Customer.name.ilike(f"%{q}%"))
                )

        try:
            page = int(request.args.get('page', 1))
        except Exception:
            page = 1
        try:
            per_page = int(request.args.get('per_page', 10))
        except Exception:
            per_page = 10

        pagination = all_sales_q.order_by(Sale.timestamp.desc()).paginate(page=page, per_page=per_page, error_out=False)
        all_sales = pagination.items

        total_all_val = db.session.query(func.coalesce(func.sum(Sale.total_price), 0.0)).select_from(all_sales_q.subquery()).scalar()
        try:
            total_all = float(total_all_val)
        except Exception:
            total_all = sum(s.total_price for s in all_sales)

        return render_template('search_sales.html', all_sales=all_sales, pagination=pagination, total_all=total_all, q=q, from_date=from_date, to_date=to_date)

    @app.route('/sales/export')
    @admin_required
    def export_sales():
        # Export sales matching the search filters to an Excel file
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            flash('The Excel export feature requires the openpyxl package. Please run `pip install openpyxl` in your virtualenv.', 'warning')
            return redirect(url_for('search_sales'))
        q = request.args.get('q', '').strip()
        from_date = request.args.get('from_date')
        to_date = request.args.get('to_date')

        all_sales_q = Sale.query
        try:
            if from_date:
                fd = datetime.strptime(from_date, '%Y-%m-%d')
                all_sales_q = all_sales_q.filter(Sale.timestamp >= fd)
            if to_date:
                td = datetime.strptime(to_date, '%Y-%m-%d')
                td_end = datetime(td.year, td.month, td.day, 23, 59, 59)
                all_sales_q = all_sales_q.filter(Sale.timestamp <= td_end)
        except Exception:
            pass

        if q:
            if q.isdigit():
                all_sales_q = all_sales_q.filter(Sale.id == int(q))
            else:
                all_sales_q = all_sales_q.join(Sale.medicine).outerjoin(Sale.customer).filter(
                    (Medicine.name.ilike(f"%{q}%")) |
                    (Customer.name.ilike(f"%{q}%"))
                )

        sales = all_sales_q.order_by(Sale.timestamp.desc()).all()

        # Build Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sales'

        headers = ['ID', 'Medicine', 'Customer', 'Quantity', 'Price per Unit', 'Total Price', 'Timestamp']
        bold = Font(bold=True)
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = bold
            cell.alignment = Alignment(horizontal='center')

        for r, s in enumerate(sales, start=2):
            ws.cell(row=r, column=1, value=s.id)
            ws.cell(row=r, column=2, value=s.medicine.name if s.medicine else '')
            ws.cell(row=r, column=3, value=s.customer.name if s.customer else '')
            ws.cell(row=r, column=4, value=s.quantity)
            pcu = ws.cell(row=r, column=5, value=s.price_per_unit)
            pcu.number_format = '#,##0.00'
            tot = ws.cell(row=r, column=6, value=s.total_price)
            tot.number_format = '#,##0.00'
            ts = ws.cell(row=r, column=7, value=s.timestamp.strftime('%Y-%m-%d %H:%M:%S') if s.timestamp else '')
            ts.alignment = Alignment(horizontal='center')

        # Simple column widths
        widths = [8, 30, 25, 10, 15, 15, 20]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        out = BytesIO()
        wb.save(out)
        out.seek(0)

        filename = f"sales_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(out, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/reports/reset/confirm', methods=['POST'])
    @admin_required
    def reset_confirm():
        period = request.form.get('period')
        if not period:
            flash('Please select a reset period.', 'danger')
            return redirect(url_for('reports'))

        now = datetime.now()
        if period == 'daily':
            cutoff = now - timedelta(hours=24)
        elif period == 'weekly':
            cutoff = now - timedelta(days=7)
        elif period == 'monthly':
            cutoff = now - timedelta(days=30)
        elif period == 'half_yearly':
            cutoff = now - timedelta(days=180)
        elif period == 'yearly':
            cutoff = now - timedelta(days=365)
        else:
            flash('Invalid period selected.', 'danger')
            return redirect(url_for('reports'))

        # Calculate what will be deleted
        sales_to_delete = Sale.query.filter(Sale.timestamp < cutoff).all()
        sales_count = len(sales_to_delete)
        total_value = sum(s.total_price for s in sales_to_delete)

        return render_template('reset_confirm.html', period=period, sales_count=sales_count, total_value=total_value, cutoff_date=cutoff.strftime('%Y-%m-%d %H:%M:%S'))

    @app.route('/reports/reset', methods=['POST'])
    @admin_required
    def reset_sales():
        period = request.form.get('period')
        if not period:
            flash('Please select a reset period.', 'danger')
            return redirect(url_for('reports'))

        now = datetime.now()
        if period == 'daily':
            cutoff = now - timedelta(hours=24)
        elif period == 'weekly':
            cutoff = now - timedelta(days=7)    
        elif period == 'monthly':
            cutoff = now - timedelta(days=30)
        elif period == 'half_yearly':
            cutoff = now - timedelta(days=180)
        elif period == 'yearly':
            cutoff = now - timedelta(days=365)
        else:
            flash('Invalid period selected.', 'danger')
            return redirect(url_for('reports'))

        # Delete sales older than cutoff
        deleted_count = Sale.query.filter(Sale.timestamp < cutoff).delete()
        db.session.commit()

        flash(f'Sales reset successfully. Deleted {deleted_count} old sales records.', 'success')
        return redirect(url_for('reports'))

    return app


# Create app instance for Gunicorn (Railway/production)
app = create_app()


if __name__ == '__main__':
    # Print all registered endpoints and their URLs for debugging
    with app.app_context():
        print('Registered endpoints:')
        for rule in app.url_map.iter_rules():
            print(f"{rule.endpoint:20} {rule.rule}")
    # Run in debug for development / MVP
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
